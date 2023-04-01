import os
import datetime
import traceback
from decimal import Decimal
import re

from openpyxl.cell import Cell
from openpyxl.cell.read_only import EmptyCell
from openpyxl.reader.excel import load_workbook
from openpyxl.styles.fills import PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from copy import copy

global global_self


# 定义计数器
class Counter:
    def __init__(self):
        self.value = True


# 设置为红色,但是不保存文件,请手动保存
def set_cell_to_red_no_save(cell):
    fill = PatternFill(fill_type='solid', fgColor="FFFF0000", bgColor="FFFF0000")
    cell.fill = copy(fill)


# 点击了开始按钮
def 循环判断文件(list_file_name):
    excel_list_file = []
    # 在这里判断那些文件需不需要执行
    for file_name in list_file_name:
        if file_name.find("国网") > -1:
            excel_list_file.append(file_name)

    print("当前文件夹内需要处理的工作表是  ", excel_list_file)
    return excel_list_file


# 获取文件绝对路径
def get_route(path, file_name):
    return path + "\\" + file_name


def start(self, path, wx):
    # 获取指定path路径下的文件列表
    list_file_name = get_list_file_by_path(wx, path)
    if list_file_name == -1:
        set_m_gauge_value(self, 0)
        return

    # 判断文件夹里文件名带国网的文件就是需要执行的文件
    excel_list_file = 循环判断文件(list_file_name)
    # 购售电异样列表
    gou_shou_dian_exception_excel_name_list = []
    # 电力销售异常列表
    power_sale_exception_excel_name_list = []
    # 而即将市场异常列表
    er_ji_exception_excel_name_list = []

    for excel in excel_list_file:
        # 在这里处理文件夹内所有的公司
        counter = Counter()
        print("当前处理第-- ", excel_list_file.index(excel) + 1, " --文件, ", "当前执行的文件是:  ", excel)
        diyibu(path, list_file_name, excel, counter)
        dierbu(path, list_file_name, excel, counter)
        disanbu(path, list_file_name, excel, counter)
        disibu(path, list_file_name, excel, counter)
        if counter.value:

            print("{:>30} {}".format(excel, "购售电处理无误"), "\n")
        else:
            gou_shou_dian_exception_excel_name_list.append(excel)
            print("{:>30} ****{}****".format(excel, "购售电有异常,请查看"), "\n")

        counter = Counter()
        counter.dian_li_xiao_shou = True
        diliubu(path, list_file_name, excel, counter)
        if counter.dian_li_xiao_shou:
            print("{:>30} {}".format(excel, "电力销售表无误"), "\n")
        else:
            power_sale_exception_excel_name_list.append(excel)
            print("{:>30} ****{}****".format(excel, "电力销售表有异常,请查看"), "\n")

        counter.er_ji_shi_chang = True
        er_ji_shi_chang(path, list_file_name, excel, counter)
        if counter.er_ji_shi_chang:
            print("{:>30} {}".format(excel, "二级市场表无误"), "\n")

        else:
            er_ji_exception_excel_name_list.append(excel)
            print("{:>30} ****{}****".format(excel, "二级市场有异常,请查看"), "\n")

    print("工作结束\n")
    set_m_gauge_value(self, 100)
    print("--  购售电结果是: 正确", len(excel_list_file) - len(gou_shou_dian_exception_excel_name_list), "   错误",
          len(gou_shou_dian_exception_excel_name_list), "个")

    if len(gou_shou_dian_exception_excel_name_list) == 0:
        print(len(excel_list_file), "个文件均无错误")
        print("")
    else:
        for excep in gou_shou_dian_exception_excel_name_list:
            print("{:>30} ****{}****".format(excep, "购售电有异常,请查看"))

        print("")
    print("-- 电力销售结果是: 正确", len(excel_list_file) - len(power_sale_exception_excel_name_list), "   错误",
          len(power_sale_exception_excel_name_list), "个")

    if len(power_sale_exception_excel_name_list) == 0:
        print(len(excel_list_file), "个文件均无错误")
        print("")
    else:
        for excep in power_sale_exception_excel_name_list:
            print("{:>30} ****{}****".format(excep, "电力销售有异常,请查看"))

        print("")
    print("-- 二级市场结果是: 正确", len(excel_list_file) - len(er_ji_exception_excel_name_list), "   错误",
          len(er_ji_exception_excel_name_list), "个")

    if len(er_ji_exception_excel_name_list) == 0:
        print(len(excel_list_file), "个文件均无错误")
        print("")
    else:
        for excep in er_ji_exception_excel_name_list:
            print("{:>30} ****{}****".format(excep, "二级市场有异常,请查看"))

    print("")


# 定义循环状态类
class RowStatus:
    def __init__(self, status):
        self.status = status

    pass


# 判断税率和单价
def judge_tax_rate_and_unit_price(sheet, row, column, data_list, counter, workbook, route):
    cell = sheet.cell(row, column)
    if cell.value is not None:
        if cell.value in data_list:
            pass
        else:
            # print(cell.value, "不在集合里")
            print("----在这里输出,判断税率和单价和预期不符合-----", cell.value)
            counter.er_ji_shi_chang = False
            set_cell_to_red_no_save(cell)


def er_ji_shi_chang(path, list_file_name, excel, counter):
    sheet, manual_table_name, workbook = get_workbook_sheet(path, list_file_name, excel, "二级市场", False,
                                                            True)
    route = get_route(path, excel)

    sheet: Worksheet
    # 拿到最大行
    max_row = sheet.max_row + 1
    # 状态标记
    row_status = RowStatus("None")
    shui_lv_list = ["1%", "3%", "13%", 0.01, 0.03, 0.13]
    dan_jia_list = [373.08, "373.08", 373.09, "373.09", 373.10, "373.10", 373.1, "373.1", 373.11, "373.11", 373.12,
                    "373.12", 386.7, "386.7"]

    # 遍历所有行
    for row in range(8, max_row):
        cell = sheet.cell(row, 2)
        cell: Cell
        value = cell.value
        if value is None:
            continue
        if value.find("水电") > -1:
            # 直接判断是不是核电,如果是就证明没有水电直接跳过
            hedian = sheet.cell(row + 1, 2)
            if hedian.value.find("4.核电") > -1:
                continue
            else:
                # 如果不是核电说明有水电
                row_status.status = "shui_dian"
                continue

        if value == "7.太阳能发电":
            row_status.status = "太阳能发电"
            continue

        if value == "8.其他能源含从公司系统外购电":
            row_status.status = "None"
            continue

        if value.find("从省级以下电网企业购电") > -1:
            # 直接判断是不是核电,如果是就证明没有水电直接跳过
            cell = sheet.cell(row + 1, 2)
            if cell.value.find("从系统内发电企业购电含抽水蓄能") > -1:
                continue
            else:
                # 如果不是核电说明有水电
                row_status.status = "从省级以下电网企业购电"
                continue
        if value.find("六、与成本费用表校验差") > -1:
            # 直接判断是不是核电,如果是就证明没有水电直接跳过
            cell = sheet.cell(row, 20)
            if cell.value is not None:
                set_cell_to_red_no_save(cell)
                counter.er_ji_shi_chang = False
                print("----六、与成本费用表校验差     不是空")

        if row_status.status == "shui_dian":
            # 检查税率 tax rate

            judge_tax_rate_and_unit_price(sheet, row, 8, shui_lv_list, counter, workbook, route)
            # 检查单价
            judge_tax_rate_and_unit_price(sheet, row, 14, dan_jia_list, counter, workbook, route)

            # 判断下一行是不是核电,如果是  row_status.status 设置为None
            he_dian = sheet.cell(row + 1, 2)
            if he_dian.value.find("4.核电") > -1:
                row_status.status = "None"
                continue

            pass

        elif row_status.status == "太阳能发电":
            if sheet.cell(row, 2).value.find("集中式光伏上网电量") > -1:
                continue
            if sheet.cell(row, 2).value.find("分布式光伏上网电量") > -1:
                continue
            if sheet.cell(row, 2).value.find("自发自用，余电上网") > -1:
                continue

            if sheet.cell(row, 2).value.find("其中：自然人") > -1:
                continue

            if sheet.cell(row, 2).value.find("非自然人") > -1:
                continue
            if sheet.cell(row, 2).value.find("全额上网") > -1:
                continue
            if sheet.cell(row, 2).value.find("其中：自然人") > -1:
                continue
            if sheet.cell(row, 2).value.find("非自然人") > -1:
                continue
            # 检查税率
            judge_tax_rate_and_unit_price(sheet, row, 8, shui_lv_list, counter, workbook, route)
            # 检查单价
            judge_tax_rate_and_unit_price(sheet, row, 14, dan_jia_list, counter, workbook, route)

            # 判断下一行是不是全额上网


        elif row_status.status == "从省级以下电网企业购电":
            # 检查税率
            judge_tax_rate_and_unit_price(sheet, row, 8, shui_lv_list, counter, workbook, route)
            cell_2 = sheet.cell(row + 1, 2)
            if cell_2.value.find("从系统内发电企业购电含抽水蓄能") > -1:
                row_status.status = "None"
                continue
            pass

    workbook.save(route)
    workbook.close()
    pass


def is_none(cell):
    if cell.value is None:
        print(f"当前单元格{cell.row}行{cell.column}列的值为空")
        cell: Cell
        cell.value = float(0)


# 第一步(本年累计)核对 购电量
def diyibu(path, list_file_name, excel, counter):
    # 获取workbook对象,目的是用来保存当前excel表格.   获取sheet对象,目的是操作某一sheet工作表
    sheet, manual_table_name, workbook = get_workbook_sheet(path, list_file_name, excel, "购售电", False,
                                                            True)
    route = get_route(path, excel)

    # cp1核对 用于公司系统内售电
    cell1 = sheet.cell(35, 5)
    cell2 = sheet.cell(39, 5)

    # cp2核对 用于省内居民农业其他用户
    cell3 = sheet.cell(25, 5)
    cell4 = sheet.cell(29, 5)
    cell5 = sheet.cell(42, 5)

    # cp3 判断 其中合计是否等于购电量合计
    cell6 = sheet.cell(21, 5)
    cell_list = [cell1, cell2, cell3, cell4, cell5, cell6]
    for cell in cell_list:
        is_none(cell)

    try:
        if Decimal(str(cell1.value)) == Decimal(str(cell2.value)):
            pass
            # print("T1用于公司系统内售电核对一致")
        else:
            set_cell_to_red_no_save(cell2)
            counter.value = False
            print("F1用于公司系统内售电核错误,请检查")
    except Exception:
        # 就执行except里的代码.
        traceback.print_exc()
        set_cell_to_red_no_save(cell2)
        counter.value = False
        print("F1用于公司系统内售电核错误,请检查")

    try:
        if Decimal(str(cell5.value)) == Decimal(str(cell3.value)) + Decimal(str(cell4.value)):
            pass
            # print("T2用于省内居民农业其他用户核对一致")
        else:
            set_cell_to_red_no_save(cell5)
            counter.value = False
            print("F2用于省内居民农业其他用户核对错误,请检查")

    except Exception:
        traceback.print_exc()
        set_cell_to_red_no_save(cell5)
        counter.value = False
        print("F2用于省内居民农业其他用户核对错误,请检查")

    try:
        if Decimal(str(cell6.value)) == Decimal(str(cell2.value)) + Decimal(str(cell5.value)):
            pass
            # print("T3购电量合计核对一致")
        else:
            set_cell_to_red_no_save(cell2)
            set_cell_to_red_no_save(cell5)
            counter.value = False
            print("F3购电量合计核对错误,请检查")
    except Exception:
        traceback.print_exc()
        set_cell_to_red_no_save(cell2)
        set_cell_to_red_no_save(cell5)
        counter.value = False
        print("F3购电量合计核对错误,请检查")
    workbook.save(route)
    workbook.close()


# 第二步(本年累计)核对 售电量
def dierbu(path, list_file_name, excel, counter):
    sheet, manual_table_name, workbook = get_workbook_sheet(path, list_file_name, excel, "购售电", False,
                                                            True)
    route = get_route(path, excel)
    cell0 = sheet.cell(45, 5)
    cell1 = sheet.cell(46, 5)
    cell2 = sheet.cell(47, 5)
    cell3 = sheet.cell(51, 5)
    cell4 = sheet.cell(52, 5)
    cell5 = sheet.cell(63, 5)
    cell6 = sheet.cell(64, 5)
    cell7 = sheet.cell(65, 5)

    cell_list = [cell0, cell1, cell2, cell3, cell4, cell5, cell6, cell7]
    for cell in cell_list:
        is_none(cell)

    try:
        if Decimal(str(cell1.value)) + Decimal(str(cell2.value)) == Decimal(str(cell5.value)) + Decimal(
                str(cell6.value)):
            pass
            # print("T4省内直接参与市场(电网代理购电)用户核对一致")
        else:
            set_cell_to_red_no_save(cell5)
            set_cell_to_red_no_save(cell6)
            counter.value = False
            print("F4省内直接参与市场(电网代理购电)用户核对错误,请检查")
    except Exception:
        traceback.print_exc()
        set_cell_to_red_no_save(cell5)
        set_cell_to_red_no_save(cell6)
        counter.value = False
        print("F4省内直接参与市场(电网代理购电)用户核对错误,请检查")

    try:
        if Decimal(str(cell7.value)) == Decimal(str(cell3.value)) + Decimal(str(cell4.value)):
            pass
            # print("T5省内居民农业其他用户核对一致")
        else:
            set_cell_to_red_no_save(cell7)
            counter.value = False
            print("F5省内居民农业其他用户核对错误,请检查")
    except Exception:
        traceback.print_exc()
        set_cell_to_red_no_save(cell7)
        counter.value = False
        print("F5省内居民农业其他用户核对错误,请检查")

    try:
        if Decimal(str(cell0.value)) == Decimal(str(cell5.value)) + Decimal(str(cell6.value)) + Decimal(
                str(cell7.value)):
            pass
            # print("T6售电量合计核对一致")
        else:
            set_cell_to_red_no_save(cell7)
            counter.value = False
            print("F6售电量合计核对错误,请检查")
    except Exception:
        traceback.print_exc()
        set_cell_to_red_no_save(cell7)
        counter.value = False
        print("F6售电量合计核对错误,请检查")
    workbook.save(route)
    workbook.close()


# 第三步(本年累计)核对 购电成本
def disanbu(path, list_file_name, excel, counter):
    sheet, manual_table_name, workbook = get_workbook_sheet(path, list_file_name, excel, "购售电", False,
                                                            True)
    route = get_route(path, excel)
    cell1 = sheet.cell(92, 5)
    cell2 = sheet.cell(93, 5)
    cell3 = sheet.cell(96, 5)

    cell_list = [cell1, cell2, cell3]
    for cell in cell_list:
        is_none(cell)

    sheet, manual_table_name, workbook = get_workbook_sheet(path, list_file_name, excel, "二级市场", False,
                                                            True)
    sheet: Worksheet

    max_row = sheet.max_row + 1
    cell4_row = 0
    for row in range(2, max_row):

        name = sheet.cell(row, 2).value

        if name is None:
            continue
        if name.count("从省级以下电网企业购电") > 0:
            cell4_row = row + 1
            break
    cell4 = sheet.cell(cell4_row, 17)
    is_none(cell4)

    print("趸售电费=", cell4.value)
    try:
        if Decimal(str(cell2.value)) == Decimal(str(cell4.value)):
            pass
            # print("T7趸售电费(含税)核对一致")
        else:
            set_cell_to_red_no_save(cell2)
            counter.value = False
            print("F7趸售电费(含税)核对错误,请检查")
    except Exception:
        traceback.print_exc()
        set_cell_to_red_no_save(cell2)
        counter.value = False
        print("F7趸售电费(含税)核对错误,请检查")

    try:
        if Decimal(str(cell1.value)) == Decimal(str(cell2.value)) + Decimal(str(cell3.value)):
            pass
            # print("T8购电费(含税)核对一致")
        else:
            set_cell_to_red_no_save(cell2)
            set_cell_to_red_no_save(cell3)
            counter.value = False
            print("F8购电费(含税)核对错误,请检查")
    except Exception:
        traceback.print_exc()
        set_cell_to_red_no_save(cell2)
        set_cell_to_red_no_save(cell3)
        counter.value = False
        print("F8购电费(含税)核对错误,请检查")
    workbook.save(route)
    workbook.close()


# 第四步(本年累计)核对 用户承担电费
def disibu(path, list_file_name, excel, counter):
    sheet, manual_table_name, workbook = get_workbook_sheet(path, list_file_name, excel, "购售电", False,
                                                            True)

    route = get_route(path, excel)
    cell0 = sheet.cell(98, 5)
    cell1 = sheet.cell(99, 5)
    cell2 = sheet.cell(101, 5)
    cell3 = sheet.cell(105, 5)
    cell4 = sheet.cell(106, 5)
    cell5 = sheet.cell(117, 5)
    cell6 = sheet.cell(119, 5)
    cell7 = sheet.cell(121, 5)

    cell_list = [cell0, cell1, cell2, cell3, cell4, cell5, cell6, cell7]
    for cell in cell_list:
        is_none(cell)

    try:
        if Decimal(str(cell1.value)) + Decimal(str(cell2.value)) == Decimal(str(cell5.value)) + Decimal(
                str(cell6.value)):
            pass
            # print("T9到户电费-省内直接参与市场(电网代理购电)用户核对一致")
        else:
            set_cell_to_red_no_save(cell5)
            set_cell_to_red_no_save(cell6)
            counter.value = False
            print("F9到户电费-省内直接参与市场(电网代理购电)用户核对错误,请检查")
    except Exception:
        traceback.print_exc()
        set_cell_to_red_no_save(cell5)
        set_cell_to_red_no_save(cell6)
        counter.value = False
        print("F9到户电费-省内直接参与市场(电网代理购电)用户核对错误,请检查")

    try:

        if Decimal(str(cell7.value)) == Decimal(str(cell3.value)) + Decimal(str(cell4.value)):
            pass
            # print("T10到户电费-省内居民农业其他用户核对一致")
        else:
            set_cell_to_red_no_save(cell7)
            counter.value = False
            print("F10到户电费-省内居民农业其他用户核对错误,请检查")
    except Exception:
        traceback.print_exc()
        set_cell_to_red_no_save(cell7)
        counter.value = False
        print("F10到户电费-省内居民农业其他用户核对错误,请检查")
    try:
        if Decimal(str(cell0.value)) == Decimal(str(cell5.value)) + Decimal(str(cell6.value)) + Decimal(
                str(cell7.value)):
            pass
            # print("T11用户承担电费合计核对一致")
        else:
            set_cell_to_red_no_save(cell5)
            set_cell_to_red_no_save(cell6)
            set_cell_to_red_no_save(cell7)
            counter.value = False
            print("F11用户承担电费合计核对错误,请检查")
    except Exception:
        traceback.print_exc()
        set_cell_to_red_no_save(cell5)
        set_cell_to_red_no_save(cell6)
        set_cell_to_red_no_save(cell7)
        counter.value = False
        print("F11用户承担电费合计核对错误,请检查")
    workbook.save(route)
    workbook.close()

    # 第五步(本年累计)核对可再生补贴
    pass


# 第六步 电力销售月报表_本月合计 _本年累计合计
def diliubu(path, list_file_name, excel, counter):
    sheet, manual_table_name, workbook = get_workbook_sheet(path, list_file_name, excel, "电力销售", False,
                                                            True)
    # 取电力销售表X10 AS10
    cell1 = sheet.cell(10, 24)
    cell2 = sheet.cell(10, 45)
    cell3 = sheet.cell(4, 2)

    # 取科目汇总表-本月
    sheet, manual_table_name, workbook = get_workbook_sheet(path, list_file_name, "科目汇总表本月", "科目汇总表查询",
                                                            False,
                                                            True)

    corporation_row = get_corporation_row(sheet, cell3.value)
    cell4 = sheet.cell(corporation_row, 6)
    if cell1.value == cell4.value:
        pass
    else:
        counter.dian_li_xiao_shou = False
        print("！！！错误：电力销售月报表本月合计与科目汇总表核对有误，请检查")

    # 当A列项目名称包含当前执行单位名称，
    # 根据当前名称确定行数row，取(row,6)
    # 判断X10==（row，6）

    # 取科目汇总表-本年累计
    sheet, manual_table_name, workbook = get_workbook_sheet(path, list_file_name, "科目汇总表本年累计",
                                                            "科目汇总表查询", False,
                                                            True)

    corporation_row = get_corporation_row(sheet, cell3.value)
    cell5 = sheet.cell(corporation_row, 6)

    # print("应收电费",cell5.value)

    if cell2.value == cell5.value:
        pass
    else:
        counter.dian_li_xiao_shou = False
        print("！！！错误：电力销售月报表本年累计合计与科目汇总表核对有误，请检查")


def get_corporation_row(sheet, corporation_name):
    max_row = sheet.max_row + 1
    for row in range(2, max_row):
        name = sheet.cell(row, 1).value
        if name is None:
            continue
        if name == "单位":
            continue
        if corporation_name.find(name) > -1:
            return row


# 第七步 电力销售月报表_本月合计 _本年累计合计

# 封装 根据文件名,sheet名,获取表格操作对象
def get_workbook_sheet(path, list_file_name, file_name, sheet_name, read_only, data_only):
    route, manual_table_name = get_file_path(path, list_file_name, file_name)
    workbook = load_workbook(route, read_only=read_only, data_only=data_only)
    sheet_name = get_sheet_name_by_workbook(workbook, sheet_name)
    sheet = workbook.get_sheet_by_name(sheet_name)
    return sheet, manual_table_name, workbook  # 获取文件路径


def get_file_path(path, file_list, name):
    manual_table_name = ""
    for file_name in file_list:
        if str(file_name).count(name) > 0:
            manual_table_name = file_name

    route = path + "\\" + manual_table_name
    return route, manual_table_name


# 根据包含的名字获取sheet名字
def get_sheet_name_by_workbook(workbook, name):
    # 查看所有工作表
    sheet_names = workbook.sheetnames
    # print("查看所有工作表", sheet_names)
    work_sheet_name = ""
    # 遍历sheet
    for i in sheet_names:
        if i.__contains__(name):
            work_sheet_name = i
    # print("输出工作sheet名字\t" + work_sheet_name)
    return work_sheet_name


def get_list_file_by_path(wx, path):
    # print("path的数据类型是:",type(path))path的数据类型是: <class 'str'>
    if len(path) == 0:
        prompt_box(wx, "提示", "未选择目录程序结束")
        return -1
    try:
        # 获取所有文件
        list_file_name = os.listdir(path)
        for file_name in list_file_name:
            if file_name.endswith(".xls"):
                prompt_box(wx, "错误", "请检查 " + file_name + " 文件格式是否正确,希望是.xlsx")
                return -1
        return list_file_name
    except OSError:
        prompt_box(wx, "提示", "路径不正确")
        return -1


# 创建提示框
def prompt_box(wx, title, news):
    # 创建提示对话框
    dlg = wx.MessageDialog(None, news, title, wx.OK)
    # 显示对话框
    dlg.ShowModal()
    # 关闭对话框
    dlg.Destroy()


# 设置进度条
def set_m_gauge_value(global_self, x):
    global_self.m_gauge_进度条.SetValue(x)
